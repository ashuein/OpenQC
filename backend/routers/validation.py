"""Validation module router -- assay validation endpoints."""

from __future__ import annotations

import hashlib
import io
import json
import uuid
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import Response
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from backend.db.database import get_db
from backend.db.validation_repository import (
    create_dataset,
    create_validation_run,
    get_dataset,
    get_validation_run,
)
from backend.engine.validation_engine import (
    calculate_intra_run_precision,
    calculate_inter_run_precision,
    calculate_linearity,
    calculate_lod,
    calculate_loq,
    evaluate_acceptance,
)
from backend.models.validation_schemas import (
    AcceptanceCriterion,
    ValidationMetric,
    ValidationResult,
    ValidationRunRequest,
    ValidationUploadResponse,
)

router = APIRouter(prefix="/validation", tags=["validation"])

VALID_TYPES = {
    "lod",
    "loq",
    "precision_intra",
    "precision_inter",
    "linearity",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_excel_to_rows(file_bytes: bytes) -> list[list]:
    """Read all rows from the active sheet of an Excel file.

    Returns a list of rows where each row is a list of cell values.
    The first row is the header.
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return []
    all_rows: list[list] = []
    for row in ws.iter_rows():
        all_rows.append([c.value for c in row])
    wb.close()
    return all_rows


# ---------------------------------------------------------------------------
# POST /validation/upload
# ---------------------------------------------------------------------------

@router.post("/upload", response_model=ValidationUploadResponse)
async def upload_validation_dataset(
    file: UploadFile = File(...),
    validation_type: str = Form(...),
    db: Session = Depends(get_db),
) -> ValidationUploadResponse:
    """Upload an Excel file containing validation data."""
    if validation_type not in VALID_TYPES:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid validation_type. Must be one of: {sorted(VALID_TYPES)}",
        )

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    file_hash = hashlib.sha256(file_bytes).hexdigest()
    rows = _parse_excel_to_rows(file_bytes)
    if len(rows) < 2:
        raise HTTPException(
            status_code=400,
            detail="File must contain at least a header row and one data row",
        )

    # Row count excludes header
    row_count = len(rows) - 1
    dataset_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)

    dataset = create_dataset(
        db,
        {
            "id": dataset_id,
            "file_name": file.filename or "unknown.xlsx",
            "file_hash": file_hash,
            "validation_type": validation_type,
            "row_count": row_count,
            "uploaded_at": now,
            "raw_data": json.dumps(rows, default=str),
        },
    )

    return ValidationUploadResponse(
        dataset_id=dataset.id,
        file_name=dataset.file_name,
        validation_type=dataset.validation_type,
        row_count=dataset.row_count,
        uploaded_at=dataset.uploaded_at,
    )


# ---------------------------------------------------------------------------
# POST /validation/run
# ---------------------------------------------------------------------------

def _build_concentration_data(rows: list[list]) -> list[dict]:
    """Build concentration_data for LOQ from raw rows.

    Expects columns: concentration, replicate_value
    """
    header = [str(h).strip().lower() if h else "" for h in rows[0]]
    conc_idx = None
    val_idx = None
    for i, h in enumerate(header):
        if "concentration" in h:
            conc_idx = i
        elif "replicate" in h or "value" in h or "measured" in h:
            val_idx = i
    if conc_idx is None or val_idx is None:
        raise HTTPException(
            status_code=400,
            detail="LOQ data must have 'concentration' and 'value/replicate' columns",
        )

    conc_map: dict[float, list[float]] = {}
    for row in rows[1:]:
        try:
            conc = float(row[conc_idx])
            val = float(row[val_idx])
        except (TypeError, ValueError, IndexError):
            continue
        conc_map.setdefault(conc, []).append(val)

    return [
        {"concentration": c, "replicates": conc_map[c]}
        for c in sorted(conc_map.keys())
    ]


def _build_replicate_values(rows: list[list]) -> list[float]:
    """Extract numeric values from first data column after header."""
    header = [str(h).strip().lower() if h else "" for h in rows[0]]
    val_idx = None
    for i, h in enumerate(header):
        if "value" in h or "replicate" in h or "ct" in h or "measured" in h:
            val_idx = i
            break
    if val_idx is None:
        val_idx = 0

    values: list[float] = []
    for row in rows[1:]:
        try:
            values.append(float(row[val_idx]))
        except (TypeError, ValueError, IndexError):
            continue
    return values


def _build_run_means(rows: list[list]) -> list[dict]:
    """Build run_means for inter-run precision.

    Expects columns: run_id, mean, date
    """
    header = [str(h).strip().lower() if h else "" for h in rows[0]]
    run_idx = date_idx = mean_idx = None
    for i, h in enumerate(header):
        if "run" in h:
            run_idx = i
        elif "mean" in h:
            mean_idx = i
        elif "date" in h:
            date_idx = i

    if run_idx is None or mean_idx is None:
        raise HTTPException(
            status_code=400,
            detail="Inter-run data must have 'run_id' and 'mean' columns",
        )

    result: list[dict] = []
    for row in rows[1:]:
        try:
            result.append({
                "run_id": str(row[run_idx]),
                "mean": float(row[mean_idx]),
                "date": str(row[date_idx]) if date_idx is not None else "",
            })
        except (TypeError, ValueError, IndexError):
            continue
    return result


def _build_linearity_levels(rows: list[list]) -> list[dict]:
    """Build levels for linearity.

    Expects columns: expected, measured
    """
    header = [str(h).strip().lower() if h else "" for h in rows[0]]
    exp_idx = meas_idx = None
    for i, h in enumerate(header):
        if "expected" in h:
            exp_idx = i
        elif "measured" in h:
            meas_idx = i

    if exp_idx is None or meas_idx is None:
        raise HTTPException(
            status_code=400,
            detail="Linearity data must have 'expected' and 'measured' columns",
        )

    levels: list[dict] = []
    for row in rows[1:]:
        try:
            levels.append({
                "expected": float(row[exp_idx]),
                "measured": float(row[meas_idx]),
            })
        except (TypeError, ValueError, IndexError):
            continue
    return levels


@router.post("/run", response_model=ValidationResult)
def run_validation(
    request: ValidationRunRequest,
    db: Session = Depends(get_db),
) -> ValidationResult:
    """Run validation calculations on a stored dataset."""
    dataset = get_dataset(db, request.dataset_id)
    if dataset is None:
        raise HTTPException(status_code=404, detail="Dataset not found")

    rows = json.loads(dataset.raw_data)
    vtype = dataset.validation_type

    # Compute raw results based on validation type
    try:
        if vtype == "lod":
            values = _build_replicate_values(rows)
            raw = calculate_lod(values)
        elif vtype == "loq":
            conc_data = _build_concentration_data(rows)
            # Default CV threshold of 20% unless overridden in criteria
            cv_threshold = 20.0
            for crit in request.acceptance_criteria:
                if crit.metric == "cv_threshold":
                    cv_threshold = crit.threshold
            raw = calculate_loq(conc_data, cv_threshold)
        elif vtype == "precision_intra":
            values = _build_replicate_values(rows)
            raw = calculate_intra_run_precision(values)
        elif vtype == "precision_inter":
            run_means = _build_run_means(rows)
            raw = calculate_inter_run_precision(run_means)
        elif vtype == "linearity":
            levels = _build_linearity_levels(rows)
            raw = calculate_linearity(levels)
        else:
            raise HTTPException(status_code=400, detail=f"Unknown type: {vtype}")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    # Build results dict for acceptance evaluation
    results_for_eval = _extract_metrics_for_eval(vtype, raw)

    # Build criteria dict
    criteria_dict = {
        c.metric: {"threshold": c.threshold, "operator": c.operator}
        for c in request.acceptance_criteria
        if c.metric != "cv_threshold"
    }

    if criteria_dict:
        acceptance = evaluate_acceptance(results_for_eval, criteria_dict)
    else:
        acceptance = {"overall_status": "pass", "details": {}}

    # Build metric list
    metrics = _build_metric_list(vtype, raw, acceptance)

    validation_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)

    create_validation_run(
        db,
        {
            "id": validation_id,
            "dataset_id": request.dataset_id,
            "overall_status": acceptance["overall_status"],
            "results_json": json.dumps(raw, default=str),
            "acceptance_criteria_json": json.dumps(
                [c.model_dump() for c in request.acceptance_criteria]
            ),
            "calculated_at": now,
        },
    )

    return ValidationResult(
        validation_id=validation_id,
        dataset_id=request.dataset_id,
        validation_type=vtype,
        overall_status=acceptance["overall_status"],
        metrics=metrics,
        raw_summary=raw,
        calculated_at=now,
    )


def _extract_metrics_for_eval(vtype: str, raw: dict) -> dict:
    """Extract flat metric->value dict for acceptance evaluation."""
    if vtype == "lod":
        return {"lod": raw["lod"]}
    elif vtype == "loq":
        return {"loq": raw["loq"]}
    elif vtype == "precision_intra":
        return {"intra_cv": raw["cv"]}
    elif vtype == "precision_inter":
        return {"inter_cv": raw["cv"]}
    elif vtype == "linearity":
        return {"r_squared": raw["r_squared"], "slope": raw["slope"]}
    return {}


def _build_metric_list(
    vtype: str, raw: dict, acceptance: dict
) -> list[ValidationMetric]:
    """Build a list of ValidationMetric from raw results and acceptance."""
    metrics: list[ValidationMetric] = []
    details = acceptance.get("details", {})

    if vtype == "lod":
        d = details.get("lod", {})
        metrics.append(ValidationMetric(
            metric="lod", value=raw["lod"],
            threshold=d.get("threshold"), operator=d.get("operator"),
            status=d.get("status", "pass"),
        ))
    elif vtype == "loq":
        d = details.get("loq", {})
        metrics.append(ValidationMetric(
            metric="loq", value=raw["loq"],
            threshold=d.get("threshold"), operator=d.get("operator"),
            status=d.get("status", "pass"),
        ))
    elif vtype == "precision_intra":
        d = details.get("intra_cv", {})
        metrics.append(ValidationMetric(
            metric="intra_cv", value=raw["cv"],
            threshold=d.get("threshold"), operator=d.get("operator"),
            status=d.get("status", "pass"),
        ))
    elif vtype == "precision_inter":
        status = "not_evaluated" if raw.get("status") == "not_evaluated" else "pass"
        d = details.get("inter_cv", {})
        if d:
            status = d.get("status", status)
        metrics.append(ValidationMetric(
            metric="inter_cv", value=raw["cv"],
            threshold=d.get("threshold"), operator=d.get("operator"),
            status=status,
        ))
    elif vtype == "linearity":
        for m in ("r_squared", "slope"):
            d = details.get(m, {})
            metrics.append(ValidationMetric(
                metric=m, value=raw[m],
                threshold=d.get("threshold"), operator=d.get("operator"),
                status=d.get("status", "pass"),
            ))

    return metrics


# ---------------------------------------------------------------------------
# GET /validation/report/{id}
# ---------------------------------------------------------------------------

@router.get("/report/{run_id}")
def get_validation_report(
    run_id: str,
    db: Session = Depends(get_db),
):
    """Return validation report as PDF (if WeasyPrint is available) or JSON fallback."""
    run = get_validation_run(db, run_id)
    if run is None:
        raise HTTPException(status_code=404, detail="Validation run not found")

    dataset = get_dataset(db, run.dataset_id)

    report_data = {
        "validation_id": run.id,
        "dataset_id": run.dataset_id,
        "file_name": dataset.file_name if dataset else None,
        "validation_type": dataset.validation_type if dataset else None,
        "overall_status": run.overall_status,
        "metrics": json.loads(run.results_json).get("metrics", [])
        if isinstance(json.loads(run.results_json), dict)
        else [],
        "raw_summary": json.loads(run.results_json),
        "acceptance_criteria": json.loads(run.acceptance_criteria_json),
        "calculated_at": run.calculated_at.isoformat(),
    }

    # Build proper metrics list from stored acceptance evaluation
    # Re-derive metrics from the stored results and acceptance criteria
    raw = json.loads(run.results_json)
    criteria_list = json.loads(run.acceptance_criteria_json)
    vtype = dataset.validation_type if dataset else None

    metrics = []
    if vtype and raw:
        metrics_for_eval = _extract_metrics_for_eval(vtype, raw)
        criteria_dict = {
            c["metric"]: {"threshold": c["threshold"], "operator": c["operator"]}
            for c in criteria_list
            if c.get("metric") != "cv_threshold"
        }
        if criteria_dict:
            from backend.engine.validation_engine import evaluate_acceptance
            acceptance = evaluate_acceptance(metrics_for_eval, criteria_dict)
        else:
            acceptance = {"overall_status": run.overall_status, "details": {}}
        metrics = [
            m.model_dump() for m in _build_metric_list(vtype, raw, acceptance)
        ]

    report_data["metrics"] = metrics

    try:
        from backend.engine.report_engine import generate_validation_report
        pdf_bytes = generate_validation_report(report_data)
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f'inline; filename="validation_report_{run_id}.pdf"'
            },
        )
    except ImportError:
        # WeasyPrint not available — fall back to JSON
        return {
            **report_data,
            "note": "PDF generation unavailable (WeasyPrint not installed). Returning JSON.",
        }
