"""Parser for QuantStudio-format .xlsx QC export files."""

from __future__ import annotations

import io

from openpyxl import load_workbook

from backend.parsers.base_parser import BaseParser


# Column names that signal a QuantStudio export
_QS_MARKER_COLUMNS = {"Well", "Well Position", "Sample Name", "Target Name", "CT"}
_QS_MARKER_COLUMNS_ALT = {"Well", "Well Position", "Sample Name", "Target Name", "Ct"}


class QuantStudioParser(BaseParser):
    """Parser for Applied Biosystems QuantStudio .xlsx exports."""

    def can_handle(self, file_metadata: dict) -> bool:
        """Return True if the file has QuantStudio-style columns."""
        filename: str = file_metadata.get("filename", "")
        if not filename.lower().endswith(".xlsx"):
            return False

        # Peek at the first row of the workbook to check column headers
        file_bytes = file_metadata.get("file_bytes")
        if file_bytes is None:
            return False

        try:
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                return False
            headers = set()
            for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                if cell.value is not None:
                    headers.add(str(cell.value).strip())
            wb.close()
            return _QS_MARKER_COLUMNS.issubset(headers) or _QS_MARKER_COLUMNS_ALT.issubset(headers)
        except Exception:
            return False

    def parse(
        self,
        file_bytes: bytes,
        mapping_config: dict | None = None,
    ) -> dict:
        """Parse QuantStudio xlsx bytes into a canonical data dict.

        Returns
        -------
        dict
            ``{"rows": [{"control_level": str, "ct_value": float,
                          "target": str, "well": str}, ...]}``
        """
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            wb.close()
            return {"rows": []}

        rows_iter = ws.iter_rows()
        header_row = next(rows_iter)
        headers = [str(c.value).strip() if c.value is not None else "" for c in header_row]

        # Build a column-index map
        col_map: dict[str, int] = {}
        for idx, h in enumerate(headers):
            col_map[h] = idx

        # Determine CT column name (may be "CT" or "Ct")
        ct_col = "CT" if "CT" in col_map else "Ct" if "Ct" in col_map else None
        target_col = "Target Name" if "Target Name" in col_map else None
        sample_col = "Sample Name" if "Sample Name" in col_map else None
        well_col = "Well" if "Well" in col_map else None

        parsed_rows: list[dict] = []
        for row in rows_iter:
            values = [c.value for c in row]

            ct_raw = values[col_map[ct_col]] if ct_col and ct_col in col_map else None
            if ct_raw is None or str(ct_raw).strip().lower() in ("", "undetermined"):
                continue

            try:
                ct_value = float(ct_raw)
            except (ValueError, TypeError):
                continue

            sample_name = ""
            if sample_col and sample_col in col_map:
                sn = values[col_map[sample_col]]
                sample_name = str(sn).strip() if sn is not None else ""

            target = ""
            if target_col and target_col in col_map:
                tn = values[col_map[target_col]]
                target = str(tn).strip() if tn is not None else ""

            well = ""
            if well_col and well_col in col_map:
                wv = values[col_map[well_col]]
                well = str(wv).strip() if wv is not None else ""

            # Derive control level from sample name heuristics
            control_level = _derive_control_level(sample_name)

            parsed_rows.append(
                {
                    "control_level": control_level,
                    "ct_value": ct_value,
                    "target": target,
                    "well": well,
                }
            )

        wb.close()
        return {"rows": parsed_rows}

    def normalize_instrument_name(self) -> str:
        return "QuantStudio"


def _derive_control_level(sample_name: str) -> str:
    """Best-effort mapping of a sample name to L1/L2/L3."""
    lower = sample_name.lower()
    for tag in ("l1", "level 1", "low", "level1"):
        if tag in lower:
            return "L1"
    for tag in ("l2", "level 2", "mid", "medium", "level2"):
        if tag in lower:
            return "L2"
    for tag in ("l3", "level 3", "high", "level3"):
        if tag in lower:
            return "L3"
    return sample_name or "Unknown"
