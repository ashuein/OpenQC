"""Generic fallback parser for .xlsx QC files with user-supplied column mapping."""

from __future__ import annotations

import io

from openpyxl import load_workbook

from backend.parsers.base_parser import BaseParser


class GenericParser(BaseParser):
    """Fallback parser that accepts any .xlsx file with a column-mapping config."""

    def can_handle(self, file_metadata: dict) -> bool:
        """Always returns True -- this is the last-resort fallback parser."""
        return True

    def parse(
        self,
        file_bytes: bytes,
        mapping_config: dict | None = None,
    ) -> dict:
        """Parse xlsx bytes using a user-provided column mapping.

        Parameters
        ----------
        mapping_config : dict | None
            Maps canonical field names to actual column headers in the file.
            Expected keys: ``"control_level"``, ``"ct_value"``.
            Optional keys: ``"target"``, ``"well"``.
            If ``None``, defaults to ``{"control_level": "Sample Name",
            "ct_value": "CT"}``.

        Returns
        -------
        dict
            ``{"rows": [{"control_level": str, "ct_value": float,
                          "target": str, "well": str}, ...]}``
        """
        if mapping_config is None:
            mapping_config = {
                "control_level": "Sample Name",
                "ct_value": "CT",
            }

        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            wb.close()
            return {"rows": []}

        rows_iter = ws.iter_rows()
        header_row = next(rows_iter)
        headers = [str(c.value).strip() if c.value is not None else "" for c in header_row]

        col_map: dict[str, int] = {}
        for idx, h in enumerate(headers):
            col_map[h] = idx

        ct_header = mapping_config.get("ct_value", "CT")
        level_header = mapping_config.get("control_level", "Sample Name")
        target_header = mapping_config.get("target", "")
        well_header = mapping_config.get("well", "")

        parsed_rows: list[dict] = []
        for row in rows_iter:
            values = [c.value for c in row]

            # CT value
            ct_raw = values[col_map[ct_header]] if ct_header in col_map else None
            if ct_raw is None or str(ct_raw).strip().lower() in ("", "undetermined"):
                continue
            try:
                ct_value = float(ct_raw)
            except (ValueError, TypeError):
                continue

            # Control level
            level_raw = values[col_map[level_header]] if level_header in col_map else None
            control_level = str(level_raw).strip() if level_raw else "Unknown"

            # Target (optional)
            target = ""
            if target_header and target_header in col_map:
                tv = values[col_map[target_header]]
                target = str(tv).strip() if tv is not None else ""

            # Well (optional)
            well = ""
            if well_header and well_header in col_map:
                wv = values[col_map[well_header]]
                well = str(wv).strip() if wv is not None else ""

            parsed_rows.append(
                {
                    "control_level": control_level,
                    "ct_value": ct_value,
                    "target": target,
                    "well": well,
                }
            )

        wb.close()
        return {"rows": parsed_rows}

    def normalize_instrument_name(self) -> str:
        return "Generic"
